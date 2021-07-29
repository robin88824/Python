from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('NewGrades.xlsx')
ws = wb.active
#read value
print(ws)
print(ws['A1'])
print(ws['A1'].value)

#change value
# ws['A2'].value = 'test'
# wb.save('NewGrades.xlsx')

#get all sheets
print(wb.sheetnames)

#get sheet thats not active
print(wb['Sheet1'])

#create new sheet
wb.create_sheet('test')
print(wb.sheetnames)


#read from row 1 to 10,column 1 to 5
for row in range(1, 11):
    for col in range(1, 5):
        #char = chr(64 + col) #gives a character represented by integer: 64 + 1 = A, 64 + 2 = B..
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

#merge cells
# ws.merge_cells('A1:D1')
# ws.unmerge_cells('A1"D1')

#insert empty row after row 7
ws.insert_rows(7)
ws.delete_rows(7)

#shift cells
ws.move_range('C1:D11', rows=0, cols=3)

wb.save('NewGrades.xlsx')